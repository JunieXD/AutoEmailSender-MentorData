from __future__ import annotations

import csv

import pytest
from openpyxl import Workbook

from mentor_data.errors import UnsafePackageError
from mentor_data.io_utils import load_yaml
from mentor_data.uploads import (
    SAFE_COLUMNS,
    extract_github_attachment,
    parse_community_package,
)

from .helpers import build_test_repository


def test_csv_package_is_parsed_and_email_is_normalized(tmp_path) -> None:
    root = build_test_repository(tmp_path)
    path = tmp_path / "safe.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(SAFE_COLUMNS)
        writer.writerow(
            [
                "示例导师",
                " MENTOR＠EXAMPLE．EDU ",
                "教授",
                "示例大学",
                "计算机学院",
                "",
                "机器学习",
                "A Paper",
                "https://cs.example.edu/faculty/mentor",
                "https://cs.example.edu/faculty/mentor",
            ]
        )
    policy = load_yaml(root / "registry" / "policy.yml")
    records = parse_community_package(path, policy)
    assert records[0]["email"] == "mentor@example.edu"


def test_xlsx_package_preserves_multiline_list_fields(tmp_path) -> None:
    root = build_test_repository(tmp_path)
    path = tmp_path / "multiline.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(SAFE_COLUMNS)
    sheet.append(
        [
            "示例导师",
            "mentor@example.edu",
            "教授",
            "示例大学",
            "计算机学院",
            "",
            "机器学习\n分布式系统",
            "Paper One; Author A; Author B\nPaper Two",
            "https://cs.example.edu/faculty/mentor",
            "https://cs.example.edu/faculty/mentor",
        ]
    )
    workbook.save(path)
    workbook.close()

    policy = load_yaml(root / "registry" / "policy.yml")
    records = parse_community_package(path, policy)

    assert records[0]["research_direction"] == "机器学习\n分布式系统"
    assert records[0]["recent_papers"] == "Paper One; Author A; Author B\nPaper Two"


def test_csv_formula_like_cell_is_rejected(tmp_path) -> None:
    root = build_test_repository(tmp_path)
    path = tmp_path / "formula.csv"
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(SAFE_COLUMNS)
        writer.writerow(
            [
                '=HYPERLINK("https://evil.example")',
                "mentor@example.edu",
                "",
                "示例大学",
                "计算机学院",
                "",
                "",
                "",
                "",
                "https://cs.example.edu/faculty/mentor",
            ]
        )
    policy = load_yaml(root / "registry" / "policy.yml")
    with pytest.raises(UnsafePackageError, match="公式"):
        parse_community_package(path, policy)


def test_xlsx_formula_and_extra_sheet_are_rejected(tmp_path) -> None:
    root = build_test_repository(tmp_path)
    policy = load_yaml(root / "registry" / "policy.yml")

    formula_path = tmp_path / "formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(SAFE_COLUMNS)
    sheet.append(
        [
            "示例导师",
            "mentor@example.edu",
            "",
            "示例大学",
            "计算机学院",
            "",
            "=1+1",
            "",
            "",
            "https://cs.example.edu/faculty/mentor",
        ]
    )
    workbook.save(formula_path)
    workbook.close()
    with pytest.raises(UnsafePackageError, match="公式"):
        parse_community_package(formula_path, policy)

    extra_sheet_path = tmp_path / "extra.xlsx"
    workbook = Workbook()
    workbook.active.append(SAFE_COLUMNS)
    workbook.create_sheet("Hidden payload")
    workbook.save(extra_sheet_path)
    workbook.close()
    with pytest.raises(UnsafePackageError, match="只能包含一个工作表"):
        parse_community_package(extra_sheet_path, policy)


def test_xlsx_declared_dimensions_are_rejected_before_unbounded_iteration(tmp_path) -> None:
    root = build_test_repository(tmp_path)
    policy = load_yaml(root / "registry" / "policy.yml")
    policy["limits"]["max_batch_rows"] = 5
    path = tmp_path / "oversized-dimension.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(SAFE_COLUMNS)
    sheet["A100000"] = "超出范围"
    workbook.save(path)
    workbook.close()

    with pytest.raises(UnsafePackageError, match="超过 5 行限制"):
        parse_community_package(path, policy)


def test_only_one_direct_github_issue_attachment_is_accepted() -> None:
    attachment = extract_github_attachment(
        "[community.xlsx](https://github.com/user-attachments/assets/"
        "123e4567-e89b-12d3-a456-426614174000)"
    )
    assert attachment.suffix == ".xlsx"

    with pytest.raises(UnsafePackageError, match="GitHub"):
        extract_github_attachment(
            "[community.xlsx](https://evil.example/user-attachments/assets/"
            "123e4567-e89b-12d3-a456-426614174000)"
        )
    with pytest.raises(UnsafePackageError, match="只能包含一个"):
        extract_github_attachment(
            "[one.xlsx](https://github.com/user-attachments/assets/"
            "123e4567-e89b-12d3-a456-426614174000)\n"
            "[two.xlsx](https://github.com/user-attachments/assets/"
            "223e4567-e89b-12d3-a456-426614174000)"
        )
