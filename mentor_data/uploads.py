from __future__ import annotations

import csv
import io
import re
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook

from .errors import UnsafePackageError
from .normalization import normalize_email, normalize_text

SAFE_COLUMNS = [
    "name",
    "email",
    "title",
    "university",
    "school",
    "department",
    "research_direction",
    "recent_papers",
    "profile_url",
    "source_url",
]
MULTILINE_COLUMNS = {"research_direction", "recent_papers"}
FORMULA_PREFIXES = ("=", "+", "-", "@")
FORBIDDEN_XLSX_PARTS = (
    "xl/externallinks/",
    "xl/embeddings/",
    "xl/activex/",
    "xl/connections.xml",
    "vbaproject.bin",
)
ATTACHMENT_LINK_PATTERN = re.compile(
    r"\[(?P<filename>[^\]\r\n]{1,255})\]\((?P<url>https://[^\s)]+)\)"
)
GITHUB_ATTACHMENT_PATH_PATTERN = re.compile(
    r"^/user-attachments/(?:assets/[0-9A-Fa-f-]{32,36}|files/[1-9][0-9]*/[^/]+)$"
)
ALLOWED_DOWNLOAD_HOSTS = {
    "github.com",
    "objects.githubusercontent.com",
    "private-user-images.githubusercontent.com",
    "user-images.githubusercontent.com",
    "github-production-user-asset-6210df.s3.amazonaws.com",
}


@dataclass(frozen=True, slots=True)
class GitHubAttachment:
    url: str
    suffix: str


@dataclass(frozen=True, slots=True)
class CommunityPackageRow:
    batch_row: int
    sheet_row: int
    record: dict[str, str]


class _NoRedirectHandler(urllib.request.HTTPRedirectHandler):
    def redirect_request(self, req, fp, code, msg, headers, newurl):  # noqa: ANN001
        return None


def _validated_download_url(value: str, *, initial: bool) -> str:
    try:
        parsed = urllib.parse.urlsplit(value)
    except ValueError as error:
        raise UnsafePackageError("附件 URL 无效") from error
    hostname = (parsed.hostname or "").lower().rstrip(".")
    if (
        parsed.scheme != "https"
        or parsed.username is not None
        or parsed.password is not None
        or parsed.port not in (None, 443)
        or hostname not in ALLOWED_DOWNLOAD_HOSTS
        or parsed.fragment
    ):
        raise UnsafePackageError("附件 URL 不属于允许的 GitHub 下载地址")
    if initial and (
        hostname != "github.com"
        or parsed.query
        or not GITHUB_ATTACHMENT_PATH_PATTERN.fullmatch(parsed.path)
    ):
        raise UnsafePackageError("附件必须是 GitHub Issue 中直接上传的文件")
    return urllib.parse.urlunsplit(parsed)


def extract_github_attachment(value: str) -> GitHubAttachment:
    matches = list(ATTACHMENT_LINK_PATTERN.finditer(value))
    if len(matches) != 1:
        raise UnsafePackageError("社区共享包字段必须且只能包含一个 GitHub 附件链接")
    match = matches[0]
    filename = match.group("filename").strip()
    if Path(filename).name != filename or any(character in filename for character in ("/", "\\")):
        raise UnsafePackageError("附件文件名无效")
    suffix = Path(filename).suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise UnsafePackageError("GitHub 附件必须是 CSV 或 XLSX")
    return GitHubAttachment(
        url=_validated_download_url(match.group("url"), initial=True),
        suffix=suffix,
    )


def download_github_attachment(
    attachment: GitHubAttachment,
    destination_directory: Path,
    *,
    max_bytes: int,
) -> Path:
    if max_bytes <= 0:
        raise UnsafePackageError("附件大小限制无效")
    opener = urllib.request.build_opener(_NoRedirectHandler())
    current_url = _validated_download_url(attachment.url, initial=True)
    response = None
    for redirect_count in range(6):
        request = urllib.request.Request(
            current_url,
            headers={
                "Accept": "application/octet-stream,*/*;q=0.1",
                "Accept-Encoding": "identity",
                "User-Agent": "AutoEmailSender-MentorData/1",
            },
            method="GET",
        )
        try:
            response = opener.open(request, timeout=20)
            break
        except urllib.error.HTTPError as error:
            if error.code not in {301, 302, 303, 307, 308}:
                raise UnsafePackageError(f"GitHub 附件下载失败：HTTP {error.code}") from error
            location = error.headers.get("Location")
            if not location or redirect_count == 5:
                raise UnsafePackageError("GitHub 附件重定向无效或次数过多") from error
            redirected = urllib.parse.urljoin(current_url, location)
            current_url = _validated_download_url(redirected, initial=False)
        except (urllib.error.URLError, TimeoutError) as error:
            raise UnsafePackageError("GitHub 附件下载失败或超时") from error
    if response is None:
        raise UnsafePackageError("GitHub 附件下载失败")

    with response:
        content_length = response.headers.get("Content-Length")
        if content_length:
            try:
                declared_size = int(content_length)
            except ValueError as error:
                raise UnsafePackageError("GitHub 附件 Content-Length 无效") from error
            if declared_size <= 0 or declared_size > max_bytes:
                raise UnsafePackageError("GitHub 附件超过大小限制")
        payload = response.read(max_bytes + 1)
    if not payload or len(payload) > max_bytes:
        raise UnsafePackageError("GitHub 附件为空或超过大小限制")
    destination_directory.mkdir(parents=True, exist_ok=True)
    destination = destination_directory / f"community-package{attachment.suffix}"
    destination.write_bytes(payload)
    return destination


def _limit(policy: dict[str, Any], key: str) -> int:
    value = policy.get("limits", {}).get(key)
    if not isinstance(value, int) or value <= 0:
        raise UnsafePackageError(f"策略缺少有效限制：{key}")
    return value


def _safe_cell(
    value: Any,
    *,
    max_characters: int,
    preserve_newlines: bool = False,
) -> str:
    if value is None:
        return ""
    raw_text = str(value)
    if preserve_newlines:
        text = "\n".join(
            normalized_line
            for line in raw_text.splitlines()
            if (normalized_line := normalize_text(line))
        )
    else:
        text = normalize_text(raw_text)
    if len(text) > max_characters:
        raise UnsafePackageError("单元格内容超过长度限制")
    if text.startswith(FORMULA_PREFIXES):
        raise UnsafePackageError("共享包不得包含公式或公式样式单元格")
    return text


def _rows_to_package_rows(
    rows: list[list[Any]],
    policy: dict[str, Any],
) -> list[CommunityPackageRow]:
    if not rows:
        raise UnsafePackageError("共享包为空")
    header = [normalize_text(str(value)) if value is not None else "" for value in rows[0]]
    if header != SAFE_COLUMNS:
        raise UnsafePackageError(f"共享包表头必须严格为：{', '.join(SAFE_COLUMNS)}")
    max_rows = _limit(policy, "max_batch_rows")
    if len(rows) - 1 > max_rows:
        raise UnsafePackageError(f"共享包超过 {max_rows} 行限制")
    max_characters = _limit(policy, "max_cell_characters")
    records: list[CommunityPackageRow] = []
    for row_number, row in enumerate(rows[1:], start=2):
        values = [
            _safe_cell(
                row[index] if index < len(row) else None,
                max_characters=max_characters,
                preserve_newlines=column in MULTILINE_COLUMNS,
            )
            for index, column in enumerate(SAFE_COLUMNS)
        ]
        if not any(values):
            continue
        record = dict(zip(SAFE_COLUMNS, values, strict=True))
        record["email"] = normalize_email(record["email"])
        records.append(
            CommunityPackageRow(
                batch_row=len(records) + 1,
                sheet_row=row_number,
                record=record,
            )
        )
    return records


def _rows_to_records(rows: list[list[Any]], policy: dict[str, Any]) -> list[dict[str, str]]:
    package_rows = _rows_to_package_rows(rows, policy)
    for item in package_rows:
        record = item.record
        if not record["name"] or not record["email"] or not record["source_url"]:
            raise UnsafePackageError(
                f"第 {item.sheet_row} 行缺少 name、email 或 source_url"
            )
    return [item.record for item in package_rows]


def _inspect_xlsx_archive(path: Path, policy: dict[str, Any]) -> None:
    max_entries = _limit(policy, "max_xlsx_entries")
    max_total = _limit(policy, "max_xlsx_uncompressed_bytes")
    max_ratio = _limit(policy, "max_xlsx_compression_ratio")
    try:
        archive = zipfile.ZipFile(path)
    except zipfile.BadZipFile as error:
        raise UnsafePackageError("XLSX 不是有效 ZIP/OOXML 文件") from error
    with archive:
        entries = archive.infolist()
        if len(entries) > max_entries:
            raise UnsafePackageError("XLSX ZIP 条目过多")
        total = 0
        for entry in entries:
            normalized_name = entry.filename.replace("\\", "/")
            pure = PurePosixPath(normalized_name)
            if pure.is_absolute() or ".." in pure.parts:
                raise UnsafePackageError("XLSX 包含路径穿越条目")
            lowered = normalized_name.lower()
            if any(part in lowered for part in FORBIDDEN_XLSX_PARTS):
                raise UnsafePackageError(f"XLSX 包含禁止的外部或嵌入内容：{entry.filename}")
            if entry.flag_bits & 0x1:
                raise UnsafePackageError("XLSX 不得包含加密 ZIP 条目")
            total += entry.file_size
            if total > max_total:
                raise UnsafePackageError("XLSX 解压后总大小超过限制")
            if entry.file_size > 0:
                if entry.compress_size == 0:
                    raise UnsafePackageError("XLSX 包含异常零压缩大小条目")
                if entry.file_size / entry.compress_size > max_ratio:
                    raise UnsafePackageError("XLSX 包含异常高压缩比条目")


def _read_community_package_rows(path: Path, policy: dict[str, Any]) -> list[list[Any]]:
    path = path.resolve()
    max_bytes = _limit(policy, "max_upload_bytes")
    size = path.stat().st_size
    if size <= 0 or size > max_bytes:
        raise UnsafePackageError(f"文件大小必须在 1 到 {max_bytes} 字节之间")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        try:
            content = path.read_bytes().decode("utf-8-sig")
        except UnicodeDecodeError as error:
            raise UnsafePackageError("CSV 必须使用 UTF-8 编码") from error
        return list(csv.reader(io.StringIO(content)))
    if suffix != ".xlsx":
        raise UnsafePackageError("只支持 CSV 或普通 XLSX 共享包")

    _inspect_xlsx_archive(path, policy)
    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    except Exception as error:  # openpyxl raises several format-specific exception classes
        raise UnsafePackageError("XLSX 无法安全读取") from error
    try:
        if len(workbook.sheetnames) != 1:
            raise UnsafePackageError("共享包必须且只能包含一个工作表")
        sheet = workbook.active
        rows: list[list[Any]] = []
        for row in sheet.iter_rows():
            values: list[Any] = []
            for cell in row:
                if cell.data_type == "f":
                    raise UnsafePackageError("XLSX 不得包含公式")
                values.append(cell.value)
            rows.append(values)
        return rows
    finally:
        workbook.close()


def parse_community_package_rows(
    path: Path,
    policy: dict[str, Any],
) -> list[CommunityPackageRow]:
    return _rows_to_package_rows(_read_community_package_rows(path, policy), policy)


def parse_community_package(path: Path, policy: dict[str, Any]) -> list[dict[str, str]]:
    return _rows_to_records(_read_community_package_rows(path, policy), policy)
